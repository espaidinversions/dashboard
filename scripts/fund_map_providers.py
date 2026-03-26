"""
fund_map_providers.py
─────────────────────
Builds a providers map for all portfolio instruments (ETFs + funds):
  - custodian: which bank/account holds it (CaixaBank, UBS, Bankinter, Abel Font…)
  - fund_house: who manages the fund (BlackRock, Vanguard, Schroders, DWS…)
  - selection: "direct" (we chose it) vs "delegated" (bank mandate)

Sources:
  - ETF sheets: "ETf's Espai RV" / "ETf's Espai RF" → Banc column (custodian)
  - Master sheet: Gestor column (custodian bank code)
  - Fund name → fund_house via pattern matching

Output:
  ../Mercats Públics/providers_map.json

Usage:
    python scripts/fund_map_providers.py

Edit the output JSON to fix any wrong entries — it is committed as source of truth.
"""

import io
import json
import os
import re
import shutil
import sys
from pathlib import Path

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Paths ─────────────────────────────────────────────────────────────────────

XLSX_GLOB   = "Mercats Públics/*.xlsx"
OUT_PATH    = Path(__file__).parent.parent / "Mercats Públics" / "providers_map.json"
ROOT        = Path(__file__).parent.parent

# ── Custodian normalisation ───────────────────────────────────────────────────
# Maps raw Gestor/Banc values → canonical custodian name

CUSTODIAN_MAP = {
    "CAIXA":      "CaixaBank",
    "CAIXA*":     "CaixaBank",       # direct-selected within CaixaBank
    "CS":         "Credit Suisse",   # acquired by UBS 2023
    "UBS":        "UBS",
    "JPM":        "JPMorgan",
    "Abel":       "Abel Font",
    "ABEL":       "Abel Font",
    "Bankinter":  "Bankinter",
    "BANKINTER":  "Bankinter",
    # ETF-sheet values that indicate custodian directly
    "WAM":        "Andbank/WAM",
    "ANDBANK":    "Andbank/WAM",
}

# ── Fund-house pattern matching ───────────────────────────────────────────────
# Each entry: (regex pattern, canonical fund house name)
# Ordered longest/most-specific first to avoid false matches.

FUND_HOUSE_PATTERNS = [
    # Multi-word first
    (r"BNY\s+Mellon",               "BNY Mellon IM"),
    (r"Brown\s+Advisory",           "Brown Advisory"),
    (r"Capital\s+Group",            "Capital Group"),
    (r"Edmond\s+de\s+Rothschild|EdR","Edmond de Rothschild AM"),
    (r"Federated\s+Hermes",         "Federated Hermes"),
    (r"Flossbach\s+von\s+Storch",   "Flossbach von Storch"),
    (r"Franklin",                   "Franklin Templeton"),
    (r"Goldman\s+Sachs|GS\s+EMERG", "Goldman Sachs AM"),
    (r"Janus\s+H[ae]nd(erson|rson)","Janus Henderson"),
    (r"JPMorgan|JPM\s",             "JPMorgan AM"),
    (r"Man\s+(Fund|Jpn|GLG)",       "Man Group"),
    (r"MFS\s+Meridian|MFS\s",       "MFS Investment Management"),
    (r"Morgan\s+Stanley",           "Morgan Stanley IM"),
    (r"Neuberger\s+Berman",         "Neuberger Berman"),
    (r"Sky\s+Harbor",               "Sky Harbor"),
    (r"T\.\s*Rowe",                 "T. Rowe Price"),
    (r"Tikehau|TIKEHAU",            "Tikehau Capital"),
    (r"Tokio\s+Marine",             "Tokio Marine AM"),
    (r"Vontobel",                   "Vontobel AM"),
    (r"Wellington|Welligton",        "Wellington Management"),
    (r"William\s+Blair",            "William Blair"),
    (r"WisdomTree",                 "WisdomTree"),
    # Single-word / acronym
    (r"iShares|ISHARES|ISHRS",      "BlackRock iShares"),
    (r"BlackRock|BLACKROCK|BSF\s|BGF\s","BlackRock"),
    (r"Vanguard|VANGUARD",          "Vanguard"),
    (r"Lyxor",                      "Amundi (ex-Lyxor)"),
    (r"Amundi|AMUNDI",              "Amundi"),
    (r"DWS",                        "DWS"),
    (r"Schroder",                   "Schroders"),
    (r"Pictet",                     "Pictet AM"),
    (r"Nordea",                     "Nordea AM"),
    (r"PIMCO",                      "PIMCO"),
    (r"Fidelity",                   "Fidelity International"),
    (r"Robecosam|ROBECOSAM",        "Robeco SAM"),
    (r"Robeco|ROBECO",              "Robeco"),
    (r"Algebris",                   "Algebris"),
    (r"Bellevue\s+BB|BB\s+Adamant", "Bellevue AM"),
    (r"DPAM",                       "Degroof Petercam AM"),
    (r"Threadneedle",               "Columbia Threadneedle"),
    (r"Lazard",                     "Lazard AM"),
    (r"Acatis",                     "Acatis"),
    (r"Polar\s+Capital|Polar\s+Global","Polar Capital"),
    (r"Aberdeen|abrdn",             "abrdn"),
    (r"AXA",                        "AXA IM"),
    (r"BNP\s+Paribas",             "BNP Paribas AM"),
    (r"Lumyna",                     "Lumyna Investments"),
    (r"Carmignac",                  "Carmignac"),
    (r"Invesco",                    "Invesco"),
    (r"Mirabaud",                   "Mirabaud AM"),
    (r"Eleva",                      "Eleva Capital"),
    (r"GAMCO",                      "GAMCO (Gabelli)"),
    (r"Mirova|MIROVA",              "Mirova (Natixis)"),
    (r"Jupiter|JUPITER",            "Jupiter AM"),
    (r"Berenberg",                  "Berenberg AM"),
    (r"Infusive",                   "Infusive AM"),
    (r"Comgest",                    "Comgest"),
    (r"Mutuafondo|Mutuactivos",     "Mutuactivos (CaixaBank)"),
    (r"ANIMA",                      "Anima AM"),
    (r"SPDR",                       "State Street SPDR"),
    (r"UBS\s+(Greater|MSCI|Fund|AM|All|Asian|China|Global|USA|US\s)",
                                    "UBS AM"),
    (r"Candriam|CANDRIAM",          "Candriam"),
    (r"Hermes",                     "Federated Hermes"),
    (r"Allianz",                    "Allianz GI"),
    (r"Variopartner|MIV\s",        "MIV / Variopartner"),
    # Abbreviations / misspellings from the Excel
    (r"Janus|JANUS",               "Janus Henderson"),       # "JANUS HEN", "Janus Hndrsn"
    (r"Ishres|Ishars",             "BlackRock iShares"),     # typos
    (r"NORD\s*1|Nord\s*1",         "Nordea AM"),             # "NORD 1 LOW DURT"
    (r"^AMU\s|AMU\s+INDEX|AMU\s+MSCI", "Amundi"),           # "AMU INDEX MSCI"
    (r"^MSS\s",                    "Morgan Stanley IM"),     # "MSS GLOBAL BRANDS"
    (r"^PIC\s",                    "Pictet AM"),             # "PIC CLN ENGY"
    (r"^SS\s+Sustainable",         "State Street"),          # "SS Sustainable Equity"
    (r"PMorgan",                   "JPMorgan AM"),           # "PMorgan Funds" (missing J)
    (r"Goldman",                   "Goldman Sachs AM"),      # "Goldman India"
    (r"FundSm[ti][ti]h|Fundsmith",  "Fundsmith"),             # "FundSmtih" typo
    (r"China Internet",            "KraneShares"),           # KWEB = KraneShares
    (r"UBS\s+(Hybrid|Conver)",     "UBS AM"),                # "UBS Hybrid", "UBS Convertible"
    (r"Ishares|ISHARES|ISHRS",     "BlackRock iShares"),     # catch-all iShares (moved up already)
]


def parse_fund_house(name: str) -> str:
    """Return best-matching fund house name or 'Unknown'."""
    if not name:
        return "Unknown"
    for pattern, house in FUND_HOUSE_PATTERNS:
        if re.search(pattern, name, re.IGNORECASE):
            return house
    return "Unknown"


def clean_isin(raw: str) -> str | None:
    m = re.search(r"([A-Z]{2}[A-Z0-9]{10})", raw.upper())
    return m.group(1) if m else None


def load_excel() -> dict:
    """
    Read the Resum Financer Excel and return
    {isin: {nom, custodian, fund_house, selection}}.
    """
    xlsx_files = list(ROOT.glob(XLSX_GLOB))
    if not xlsx_files:
        print("ERROR: no .xlsx found in Mercats Públics/")
        sys.exit(1)

    tmp = Path(os.environ.get("TEMP", "C:/tmp")) / "espai_providers_tmp.xlsx"
    shutil.copy2(xlsx_files[0], tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)

    records = {}   # isin → {nom, custodian, fund_house, selection}

    # ── Master sheet ──────────────────────────────────────────────────────────
    ws = wb["Master"]
    # Find Gestor column index
    gestor_col = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row[1] == "Tipus":
            gestor_col = next((i for i, v in enumerate(row) if v == "Gestor"), None)
            break

    CLOSED = {"TANCAT", "Tancada", "Tancat", "TANCADA", "Estruct", "Autocancel", "Autocancel "}

    for row in ws.iter_rows(min_row=4, values_only=True):
        tipus = str(row[1]).strip() if row[1] is not None else ""
        nom   = str(row[2]).strip() if row[2] else ""
        isin_raw = row[5]
        gestor_raw = row[gestor_col] if gestor_col and len(row) > gestor_col else None

        if not isin_raw or not isinstance(isin_raw, str):
            continue
        isin = clean_isin(isin_raw)
        if not isin:
            continue

        # Skip section subtotal rows (no meaningful gestor string)
        if not isinstance(gestor_raw, str) or not gestor_raw.strip():
            continue

        custodian = CUSTODIAN_MAP.get(gestor_raw.strip(), gestor_raw.strip())
        is_closed = tipus in CLOSED

        if isin not in records:
            records[isin] = {
                "nom":       nom,
                "custodian": custodian,
                "fund_house": parse_fund_house(nom),
                "selection": "delegated",   # bank-managed mandate by default
                "closed":    is_closed,
            }
        else:
            # If the same ISIN appears multiple times (multiple tranches), keep
            # the custodian of the first active occurrence
            if not is_closed and records[isin].get("closed"):
                records[isin]["custodian"] = custodian
                records[isin]["closed"] = False

    # ── ETF RV sheet ──────────────────────────────────────────────────────────
    # Banc col = 0 (may be None = Abel Font direct, or "Bankinter")
    ws_rv = wb["ETf's Espai RV"]
    for row in ws_rv.iter_rows(min_row=3, values_only=True):
        banc     = str(row[0]).strip() if row[0] else None
        isin_raw = row[6]
        nom      = str(row[3]).strip() if row[3] else ""

        if not isin_raw or not isinstance(isin_raw, str):
            continue
        isin = clean_isin(isin_raw)
        if not isin:
            continue

        if banc and banc.upper() in ("BANKINTER",):
            custodian  = "Bankinter"
            selection  = "direct"
        else:
            custodian  = "Abel Font"
            selection  = "direct"

        if isin not in records:
            records[isin] = {
                "nom":        nom,
                "custodian":  custodian,
                "fund_house": parse_fund_house(nom),
                "selection":  selection,
                "closed":     False,
            }

    # ── ETF RF sheet ──────────────────────────────────────────────────────────
    ws_rf = wb["ETf's Espai RF"]
    for row in ws_rf.iter_rows(min_row=3, values_only=True):
        isin_raw = row[6]
        nom      = str(row[2]).strip() if row[2] else ""

        if not isin_raw or not isinstance(isin_raw, str):
            continue
        isin = clean_isin(isin_raw)
        if not isin:
            continue

        if isin not in records:
            records[isin] = {
                "nom":        nom,
                "custodian":  "Abel Font",
                "fund_house": parse_fund_house(nom),
                "selection":  "direct",
                "closed":     False,
            }

    return records


def main():
    records = load_excel()

    # ── Stats ──────────────────────────────────────────────────────────────────
    from collections import Counter
    custodians = Counter(v["custodian"] for v in records.values())
    houses     = Counter(v["fund_house"] for v in records.values() if not v.get("closed"))

    print(f"Total ISINs mapped: {len(records)}")
    print()
    print("Custodians:")
    for k, n in custodians.most_common():
        print(f"  {k:<22} {n:3d} positions")
    print()
    print("Top fund houses (active only):")
    for k, n in houses.most_common(20):
        print(f"  {k:<35} {n:3d}")

    unknown = [isin for isin, v in records.items() if v["fund_house"] == "Unknown" and not v.get("closed")]
    if unknown:
        print(f"\nUnknown fund house ({len(unknown)}):")
        for isin in unknown:
            print(f"  {isin}  {records[isin]['nom'][:60]}")

    # ── Save ───────────────────────────────────────────────────────────────────
    payload = {
        "_note": "Manually review and edit where needed. Committed as source of truth.",
        "_legend": {
            "custodian": "Bank/account that holds the position",
            "fund_house": "Asset manager that runs the fund",
            "selection": "'direct' = we chose it; 'delegated' = bank mandate chose it",
            "closed": "True if position is no longer active"
        },
        "map": dict(sorted(records.items())),
    }
    OUT_PATH.parent.mkdir(exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nSaved -> {OUT_PATH}")


if __name__ == "__main__":
    main()
