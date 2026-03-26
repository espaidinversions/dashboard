"""
portfolio_build_values.py
─────────────────────────
Builds a time-series of portfolio values for every fund / ETF position.

Logic:
  value(date) = Σ tranche_i.n_titols × NAV(date)
                for all tranches where date ≥ tranche.purchase_date

Sources:
  - ETF sheets ("ETf's Espai RV/RF"): clean purchase dates, per-tranche units
  - Master sheet: bank-managed fund tranches, messier dates (uses year-end
    value columns to infer approximate start when purchase date is missing)
  - NAV prices: ../Mercats Públics/prices/<ISIN>.csv  (ETFs, from justETF)
                ../Mercats Públics/fund_prices/<ISIN>.csv  (funds, from Morningstar)

Output:
  ../Mercats Públics/portfolio_value.csv
      date, isin, nom, custodian, units, nav, value_eur

Usage:
    python scripts/portfolio_build_values.py
"""

import io
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT          = Path(__file__).parent.parent
PRICES_DIR    = ROOT / "Mercats Públics" / "prices"
FUND_DIR      = ROOT / "Mercats Públics" / "fund_prices"
OUT_CSV       = ROOT / "Mercats Públics" / "portfolio_value.csv"
PROVIDERS_MAP = ROOT / "Mercats Públics" / "providers_map.json"

# Year-end anchor dates (col index → date)
YEAREND_COLS = {
    27: date(2018, 12, 31),
    26: date(2019, 12, 31),
    25: date(2020, 12, 31),
    24: date(2021, 12, 31),
    23: date(2022, 12, 31),
    22: date(2023, 12, 31),
    21: date(2024, 12, 31),
    20: date(2025, 12, 31),
}

CLOSED_LABELS = {"TANCAT", "Tancada", "Tancat", "TANCADA", "Estruct", "Autocancel", "Autocancel "}

# Year the TANCATS row appears under → last date we held it (Dec 31 of that year)
TANCATS_YEAR_RE = re.compile(r"TANCATS\s+(\d{4})", re.IGNORECASE)

CUSTODIAN_CODES = {
    "CAIXA": "CaixaBank", "CAIXA*": "CaixaBank",
    "CS": "Credit Suisse", "UBS": "UBS",
    "JPM": "JPMorgan", "Abel": "Abel Font", "ABEL": "Abel Font",
    "Bankinter": "Bankinter", "BANKINTER": "Bankinter",
}


def clean_isin(raw: str) -> str | None:
    m = re.search(r"([A-Z]{2}[A-Z0-9]{10})", raw.upper())
    return m.group(1) if m else None


def parse_date(val) -> date | None:
    """Convert various Excel date formats to a date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # Excel serial date (1 = 1900-01-01 with Excel's off-by-one)
        if val > 1000:   # plausible Excel serial
            try:
                return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
            except Exception:
                pass
    return None


def load_tranches_etf(wb) -> list[dict]:
    """Read per-tranche data from ETF RV and RF sheets."""
    tranches = []
    for sheet_name, isin_col, nom_col, date_col, units_col in [
        ("ETf's Espai RV", 6, 3, 4, 8),
        ("ETf's Espai RF", 6, 2, 3, 8),
    ]:
        ws = wb[sheet_name]
        # Determine custodian col (col 0 in RV = bank, RF doesn't have it)
        banc_col = 0 if sheet_name == "ETf's Espai RV" else None

        for row in ws.iter_rows(min_row=3, values_only=True):
            isin_raw = row[isin_col]
            if not isin_raw or not isinstance(isin_raw, str):
                continue
            isin = clean_isin(isin_raw)
            if not isin:
                continue

            nom        = str(row[nom_col]).strip() if row[nom_col] else ""
            pdate      = parse_date(row[date_col])
            n_titols   = row[units_col]
            banc_raw   = str(row[banc_col]).strip() if banc_col is not None and row[banc_col] else None
            custodian  = CUSTODIAN_CODES.get(banc_raw, "Bankinter") if banc_raw else "Abel Font"

            if not n_titols or not isinstance(n_titols, (int, float)) or n_titols <= 0:
                continue

            tranches.append({
                "isin":       isin,
                "nom":        nom,
                "custodian":  custodian,
                "purchase_date": pdate,
                "n_titols":   float(n_titols),
                "source_sheet": sheet_name,
            })
    return tranches


def load_tranches_master(wb, include_closed: bool = False) -> list[dict]:
    """Read per-tranche data from Master sheet.

    When include_closed=True, also reads TANCATS rows and attaches an
    end_date = Dec 31 of the TANCATS year so the value series stops then.
    """
    ws = wb["Master"]

    # Find header row
    gestor_col = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row[1] == "Tipus":
            gestor_col = next((i for i, v in enumerate(row) if v == "Gestor"), None)
            break

    tranches = []
    current_tancats_year = None   # tracked while scanning rows

    for row in ws.iter_rows(min_row=4, values_only=True):
        row_str = " ".join(str(c) for c in row if c is not None)

        # Track which TANCATS year we're inside
        m = TANCATS_YEAR_RE.search(row_str)
        if m:
            current_tancats_year = int(m.group(1))

        tipus      = str(row[1]).strip() if row[1] else ""
        nom        = str(row[2]).strip() if row[2] else ""
        isin_raw   = row[5]
        n_titols   = row[9]
        gestor_raw = row[gestor_col] if gestor_col and len(row) > gestor_col else None

        # A row is closed if we are inside a TANCATS section (tracked via header rows)
        is_closed = current_tancats_year is not None

        if is_closed and not include_closed:
            continue
        if not is_closed and include_closed:
            continue  # skip active rows when collecting closed-only
        if not isin_raw or not isinstance(isin_raw, str):
            continue
        isin = clean_isin(isin_raw)
        if not isin:
            continue
        if not isinstance(gestor_raw, str) or not gestor_raw.strip():
            continue
        if not n_titols or not isinstance(n_titols, (int, float)) or n_titols <= 0:
            continue

        custodian  = CUSTODIAN_CODES.get(gestor_raw.strip(), gestor_raw.strip())
        pdate      = parse_date(row[3])

        # Infer approximate purchase date from earliest year-end value if date missing
        if pdate is None:
            for col_idx in sorted(YEAREND_COLS.keys(), reverse=True):  # oldest first
                if col_idx < len(row) and isinstance(row[col_idx], (int, float)) and row[col_idx] > 0:
                    pdate = YEAREND_COLS[col_idx]
                    break

        # For closed positions, cap the series at Dec 31 of the TANCATS year
        end_date = date(current_tancats_year, 12, 31) if is_closed else None

        tranches.append({
            "isin":          isin,
            "nom":           nom,
            "custodian":     custodian,
            "purchase_date": pdate,
            "end_date":      end_date,
            "n_titols":      float(n_titols),
            "source_sheet":  "Master",
            "closed":        is_closed,
        })
    return tranches


def load_nav_series(isin: str) -> pd.Series | None:
    """Load daily NAV/price series for an ISIN. Returns pd.Series(index=date, values=price) or None."""
    for d in [PRICES_DIR, FUND_DIR]:
        p = d / f"{isin}.csv"
        if p.exists():
            df = pd.read_csv(p, parse_dates=["date"])
            df = df.dropna(subset=["close"]).set_index("date")["close"].sort_index()
            return df
    return None


def build_portfolio_values(tranches: list[dict]) -> pd.DataFrame:
    """
    For each (isin, custodian) group: accumulate units tranche by tranche,
    then multiply by daily NAV → value time series.
    """
    # Group tranches by (isin, custodian)
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for t in tranches:
        groups[(t["isin"], t["custodian"])].append(t)

    all_rows = []

    for (isin, custodian), tranche_list in sorted(groups.items()):
        nav = load_nav_series(isin)
        if nav is None:
            print(f"  SKIP  {isin} ({custodian}) — no price data")
            continue

        nom = tranche_list[0]["nom"]

        # Sort tranches by purchase date (None last)
        tranche_list.sort(key=lambda t: t["purchase_date"] or date(2099, 1, 1))

        # For each date in NAV series, sum up units from all tranches purchased ≤ date
        nav_dates = nav.index.date

        rows = []
        for nav_dt, price in nav.items():
            nav_date = nav_dt.date() if hasattr(nav_dt, "date") else nav_dt
            units = sum(
                t["n_titols"]
                for t in tranche_list
                if t["purchase_date"] is not None and t["purchase_date"] <= nav_date
                and (t.get("end_date") is None or nav_date <= t["end_date"])
            )
            if units > 0:
                rows.append({
                    "date":       nav_date,
                    "isin":       isin,
                    "nom":        nom,
                    "custodian":  custodian,
                    "units":      units,
                    "nav":        round(price, 4),
                    "value_eur":  round(units * price, 2),
                })

        if rows:
            all_rows.extend(rows)
            start = rows[0]["date"]
            end   = rows[-1]["date"]
            last_val = rows[-1]["value_eur"]
            total_units = sum(t["n_titols"] for t in tranche_list if t["purchase_date"] is not None)
            print(f"  OK    {isin} | {custodian:<22} | {len(rows):5d} rows | "
                  f"{start} → {end} | units={total_units:,.0f} | last={last_val:>12,.0f} €")
        else:
            print(f"  EMPTY {isin} ({custodian}) — no dated tranches or no NAV overlap")

    return pd.DataFrame(all_rows)


def main():
    import openpyxl
    xlsx_files = list(ROOT.glob("Mercats Públics/*.xlsx"))
    if not xlsx_files:
        print("ERROR: no .xlsx in Mercats Públics/")
        sys.exit(1)

    tmp = Path(os.environ.get("TEMP", "C:/tmp")) / "espai_pv_tmp.xlsx"
    shutil.copy2(xlsx_files[0], tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)

    print("Loading tranches from Excel…")
    etf_tranches    = load_tranches_etf(wb)
    master_active   = load_tranches_master(wb, include_closed=False)
    master_closed   = load_tranches_master(wb, include_closed=True)
    # master_closed includes both active and closed rows; keep only closed ones
    master_closed   = [t for t in master_closed if t.get("closed")]
    all_tranches    = etf_tranches + master_active + master_closed

    print(f"  ETF sheet tranches:     {len(etf_tranches)}")
    print(f"  Master active tranches: {len(master_active)}")
    print(f"  Master closed tranches: {len(master_closed)}")
    print(f"  Total:                  {len(all_tranches)}")
    print()

    print("Building portfolio value time series…")
    df = build_portfolio_values(all_tranches)

    if df.empty:
        print("No data produced.")
        return

    df = df.sort_values(["isin", "custodian", "date"])
    df.to_csv(OUT_CSV, index=False)

    print(f"\n{'─'*64}")
    print(f"Positions processed: {df.groupby(['isin','custodian']).ngroups}")
    print(f"Total rows:          {len(df):,}")
    print(f"Date range:          {df['date'].min()} → {df['date'].max()}")
    latest = df.groupby(["isin","custodian"])["date"].max()
    last_day = df[df["date"] == df["date"].max()]
    total = last_day["value_eur"].sum()
    print(f"Latest total value:  {total:>15,.0f} €")
    print(f"Output:              {OUT_CSV}")


if __name__ == "__main__":
    main()
