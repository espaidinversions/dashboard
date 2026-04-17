"""
transactions_export_js.py
──────────────────────────
Builds a transaction log (buys + sells) for all PM positions from the Excel.

Sources:
  - ETf's Espai RV / ETf's Espai RF  → ETF buy tranches (exact date)
  - Master sheet (active)             → managed-fund buy tranches (date or year-end estimate)
  - Master sheet (TANCATS sections)   → managed-fund sells (approximated Dec 31 of TANCATS year)
  - raw-data/bank-movements-40510.json
      → exact UBS / Credit Suisse buys, onboarding transfer-ins, and sells from bank PDF

Output:
  src/generated/publicMarkets/pmTransactions.js   export const PM_TRANSACTIONS = [...]

Each entry:
  { id, action, date, isin, nom, tipus, custodian, units, nav, valueEur }

Usage:
    python -m scripts.transactions_export_js
"""

import csv
import io
import json
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from scripts.pm_model_types import PMBankMovement, PMMasterPosition, PMTransactionDraft, PMTransactionRow

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT   = Path(__file__).parent.parent
OUT_JS = ROOT / "src" / "generated" / "publicMarkets" / "pmTransactions.js"
BANK_MOVEMENTS_JSON = ROOT / "raw-data" / "bank-movements-40510.json"

YEAREND_COLS = {
    27: date(2018, 12, 31), 26: date(2019, 12, 31), 25: date(2020, 12, 31),
    24: date(2021, 12, 31), 23: date(2022, 12, 31), 22: date(2023, 12, 31),
    21: date(2024, 12, 31), 20: date(2025, 12, 31),
}

CUSTODIAN_CODES = {
    "CAIXA": "CaixaBank", "CAIXA*": "CaixaBank",
    "CS": "Credit Suisse", "UBS": "UBS",
    "JPM": "JPMorgan", "Abel": "CaixaBank", "ABEL": "CaixaBank",
    "Bankinter": "Bankinter", "BANKINTER": "Bankinter",
}

ISIN_RE      = re.compile(r"([A-Z]{2}[A-Z0-9]{10})")
TANCATS_RE   = re.compile(r"TANCATS\s+(\d{4})", re.IGNORECASE)
TIPUS_MAP    = {"RV": "RV", "RF": "RF", "Estruct": "Estruct"}
PRICES_DIR   = ROOT / "Mercats Públics" / "prices"
BANK_FAMILY  = {"UBS", "Credit Suisse"}


def classify_abel_etf_custodian(raw_banc: str | None) -> str | None:
    banc = str(raw_banc or "").strip()
    if not banc:
        return None
    if banc == "BANKINTER":
        return "Bankinter"
    if banc == "Bankinter":
        return "Interactive Brokers"
    if banc.upper() == "BANKINTER":
        return "Bankinter"
    return None


def load_etf_prices() -> dict[str, list[tuple[date, float]]]:
    """Load all ETF price CSVs → { isin: [(date, close), ...] sorted ascending }."""
    prices: dict[str, list[tuple[date, float]]] = {}
    for csv_path in PRICES_DIR.glob("*.csv"):
        isin = csv_path.stem
        rows = []
        with open(csv_path, encoding="utf-8", newline="") as f:
            for r in csv.DictReader(f):
                try:
                    rows.append((date.fromisoformat(r["date"]), float(r["close"])))
                except (KeyError, ValueError):
                    pass
        if rows:
            prices[isin] = sorted(rows)
    return prices


def find_closest_price(prices: list[tuple[date, float]], tx_date: date) -> float | None:
    """Binary-search for the closest price on or before tx_date; fall back to earliest if none."""
    if not prices:
        return None
    lo, hi = 0, len(prices) - 1
    result = None
    while lo <= hi:
        mid = (lo + hi) // 2
        if prices[mid][0] <= tx_date:
            result = prices[mid][1]
            lo = mid + 1
        else:
            hi = mid - 1
    # If no price on/before tx_date, return earliest available
    return result if result is not None else prices[0][1]


def clean_isin(v: str | None) -> str | None:
    if not v: return None
    m = ISIN_RE.search(str(v).upper())
    return m.group(1) if m else None


def parse_date(val) -> date | None:
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, (int, float)) and val > 1000:
        try: return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
        except (OverflowError, ValueError, TypeError): return None
    return None


def infer_date_from_yearend(row: list[object]) -> date | None:
    """Return earliest year-end where the position had a positive value."""
    for col_idx in sorted(YEAREND_COLS.keys(), reverse=True):
        if col_idx < len(row) and isinstance(row[col_idx], (int, float)) and row[col_idx] > 0:
            return YEAREND_COLS[col_idx]
    return None


def load_wb():
    xlsx_files = list(ROOT.glob("Mercats Públics/*.xlsx"))
    if not xlsx_files:
        print("ERROR: no .xlsx in Mercats Públics/"); sys.exit(1)
    tmp = Path(os.environ.get("TEMP", tempfile.gettempdir())) / "espai_tx_tmp.xlsx"
    shutil.copy2(xlsx_files[0], tmp)
    return openpyxl.load_workbook(tmp, data_only=True)


def normalize_bank_movement(row: object) -> PMBankMovement | None:
    if not isinstance(row, dict):
        return None
    required_keys = ("action", "section", "date", "reference", "securityCode", "nameRaw", "currency", "priceCurrency")
    if any(not isinstance(row.get(key), str) for key in required_keys):
        return None
    normalized: PMBankMovement = {
        "action": row["action"],
        "section": row["section"],
        "date": row["date"],
        "reference": row["reference"],
        "securityCode": row["securityCode"],
        "isin": row["isin"] if isinstance(row.get("isin"), str) else None,
        "nameRaw": row["nameRaw"].strip(),
        "currency": row["currency"],
        "priceCurrency": row["priceCurrency"],
        "price": float(row["price"]) if isinstance(row.get("price"), (int, float)) else None,
        "units": float(row["units"]) if isinstance(row.get("units"), (int, float)) else None,
        "valueEur": float(row["valueEur"]) if isinstance(row.get("valueEur"), (int, float)) else None,
    }
    if isinstance(row.get("tax"), (int, float)):
        normalized["tax"] = float(row["tax"])
    if isinstance(row.get("taxCurrency"), str):
        normalized["taxCurrency"] = row["taxCurrency"]
    if isinstance(row.get("grossAmount"), (int, float)):
        normalized["grossAmount"] = float(row["grossAmount"])
    if row.get("isInitialTransferIn") is True:
        normalized["isInitialTransferIn"] = True
    return normalized


def load_bank_movements() -> list[PMBankMovement]:
    if not BANK_MOVEMENTS_JSON.exists():
        return []
    try:
        payload: dict[str, object] = json.loads(BANK_MOVEMENTS_JSON.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"WARNING: could not read {BANK_MOVEMENTS_JSON}: {exc}")
        return []
    rows = payload.get("movements")
    if not isinstance(rows, list):
        return []
    return [row for row in (normalize_bank_movement(row) for row in rows) if row is not None]


def etf_buys(wb, etf_prices: dict[str, list[tuple[date, float]]]) -> list[PMTransactionDraft]:
    txs: list[PMTransactionDraft] = []
    for sheet, isin_col, nom_col, date_col, units_col, banc_col, tipus in [
        ("ETf's Espai RV", 6, 3, 4, 8, 0, "RV"),
        ("ETf's Espai RF", 6, 2, 3, 8, None, "RF"),
    ]:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            isin = clean_isin(row[isin_col]) if row[isin_col] else None
            if not isin: continue
            nom     = str(row[nom_col]).strip() if row[nom_col] else ""
            pdate   = parse_date(row[date_col])
            units   = row[units_col]
            banc    = str(row[banc_col]).strip() if banc_col is not None and row[banc_col] else None
            custodian = classify_abel_etf_custodian(banc) or "CaixaBank"
            if not units or not isinstance(units, (int, float)) or units <= 0: continue
            nav = None
            value_eur = None
            if pdate and isin in etf_prices:
                nav = find_closest_price(etf_prices[isin], pdate)
                if nav is not None:
                    value_eur = round(float(units) * nav, 0)
            txs.append({
                "action": "buy", "date": pdate, "isin": isin, "nom": nom,
                "tipus": tipus, "custodian": custodian, "units": float(units),
                "nav": nav, "valueEur": value_eur,
            })
    return txs


def master_positions(wb) -> list[PMMasterPosition]:
    """Return normalized Master rows used for fallback txs and bank-movement matching."""
    ws = wb["Master"]
    # Find Gestor column
    gestor_col = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row[1] == "Tipus":
            gestor_col = next((i for i, v in enumerate(row) if v == "Gestor"), None)
            break

    positions: list[PMMasterPosition] = []
    current_tancats_year = None

    for row in ws.iter_rows(min_row=4, values_only=True):
        row_str = " ".join(str(c) for c in row if c is not None)
        m = TANCATS_RE.search(row_str)
        if m:
            current_tancats_year = int(m.group(1))

        tipus_raw  = str(row[1]).strip() if row[1] else ""
        nom        = str(row[2]).strip() if row[2] else ""
        isin_raw   = row[5]
        n_titols   = row[9]
        nav_raw    = row[10]   # current NAV column
        value_raw  = row[11]   # current value column
        gestor_raw = row[gestor_col] if gestor_col and len(row) > gestor_col else None

        isin = clean_isin(isin_raw) if isin_raw else None
        if not isin or not nom or len(nom) < 3: continue
        if not isinstance(gestor_raw, str) or not gestor_raw.strip(): continue
        if not n_titols or not isinstance(n_titols, (int, float)) or n_titols <= 0: continue

        custodian = CUSTODIAN_CODES.get(gestor_raw.strip(), gestor_raw.strip())
        tipus     = "RV" if "RV" in tipus_raw.upper() else ("RF" if "RF" in tipus_raw.upper() else "RV")
        nav       = float(nav_raw) if isinstance(nav_raw, (int, float)) and nav_raw > 0 else None
        value_eur = float(value_raw) if isinstance(value_raw, (int, float)) and value_raw > 0 else None

        is_closed = current_tancats_year is not None

        positions.append({
            "is_closed": is_closed,
            "tancats_year": current_tancats_year if is_closed else None,
            "buy_date": parse_date(row[3]) or infer_date_from_yearend(row),
            "sell_date": date(current_tancats_year, 12, 31) if is_closed else None,
            "isin": isin,
            "nom": nom,
            "tipus": tipus,
            "custodian": custodian,
            "units": float(n_titols),
            "nav": nav,
            "valueEur": value_eur,
        })

    return positions


def approx_same_units(a: float | int | None, b: float | int | None, tolerance: float = 0.05) -> bool:
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tolerance


def choose_position_for_bank_tx(positions: list[PMMasterPosition], movement: PMBankMovement) -> PMMasterPosition | None:
    if not positions:
        return None

    tx_date = date.fromisoformat(movement["date"])
    tx_units = movement.get("units")
    action = movement["action"]

    active = [p for p in positions if not p["is_closed"]]
    closed = [p for p in positions if p["is_closed"]]
    exact = [p for p in positions if approx_same_units(p.get("units"), tx_units)]

    def sort_sell_candidates(rows):
        def key(row):
            sell_date = row.get("sell_date")
            date_distance = abs((sell_date - tx_date).days) if sell_date else 99_999
            unit_distance = abs((row.get("units") or 0) - (tx_units or 0))
            return (date_distance, unit_distance)
        return sorted(rows, key=key)

    if action == "sell":
        exact_closed = [p for p in exact if p["is_closed"]]
        if exact_closed:
            return sort_sell_candidates(exact_closed)[0]
        if exact:
            return sort_sell_candidates(exact)[0]
        if closed:
            return sort_sell_candidates(closed)[0]
        if active:
            return sort_sell_candidates(active)[0]
        return sort_sell_candidates(positions)[0]

    exact_active = [p for p in exact if not p["is_closed"]]
    if len(active) == 1:
        return active[0]
    if exact_active:
        return sorted(exact_active, key=lambda row: abs((row.get("units") or 0) - (tx_units or 0)))[0]
    if active:
        return sorted(active, key=lambda row: (row.get("buy_date") is None, row.get("buy_date") or date.max))[0]
    if exact:
        return exact[0]
    return sorted(positions, key=lambda row: (row.get("sell_date") is None, row.get("sell_date") or date.max))[0]


def bank_txs(
    master_positions_rows: list[PMMasterPosition],
    bank_movements: list[PMBankMovement],
) -> tuple[list[PMTransactionDraft], set[str], set[str]]:
    family_positions: dict[str, list[PMMasterPosition]] = {}
    for row in master_positions_rows:
        if row["custodian"] not in BANK_FAMILY:
            continue
        family_positions.setdefault(row["isin"], []).append(row)

    txs: list[PMTransactionDraft] = []
    covered_buy_isins: set[str] = set()
    covered_sell_isins: set[str] = set()

    for movement in bank_movements:
        isin = clean_isin(movement.get("isin"))
        if not isin:
            continue
        if movement.get("action") == "buy":
            tx_action = "buy"
        elif movement.get("action") == "transfer_in" and movement.get("isInitialTransferIn"):
            tx_action = "buy"
        elif movement.get("action") == "sell":
            tx_action = "sell"
        else:
            continue

        positions = family_positions.get(isin)
        if not positions:
            continue
        movement_for_match: PMBankMovement = {**movement, "action": tx_action}
        matched = choose_position_for_bank_tx(positions, movement_for_match)
        if not matched:
            continue

        nav = movement.get("price")
        value_eur = movement.get("valueEur")
        units = movement.get("units")
        if nav is None and value_eur is not None and units not in (None, 0):
            nav = float(value_eur) / float(units)

        txs.append({
            "action": tx_action,
            "date": date.fromisoformat(str(movement["date"])[:10]),
            "isin": isin,
            "nom": matched["nom"],
            "tipus": matched["tipus"],
            "custodian": matched["custodian"],
            "units": float(units) if isinstance(units, (int, float)) else None,
            "nav": float(nav) if isinstance(nav, (int, float)) else None,
            "valueEur": float(value_eur) if isinstance(value_eur, (int, float)) else None,
        })
        if tx_action == "buy":
            covered_buy_isins.add(isin)
        else:
            covered_sell_isins.add(isin)

    return txs, covered_buy_isins, covered_sell_isins


def fallback_master_txs(
    master_positions_rows: list[PMMasterPosition],
    covered_buy_isins: set[str],
    covered_sell_isins: set[str],
) -> list[PMTransactionDraft]:
    txs: list[PMTransactionDraft] = []
    for row in master_positions_rows:
        if row["custodian"] in BANK_FAMILY:
            if not row["is_closed"] and row["isin"] in covered_buy_isins:
                continue
            if row["is_closed"] and row["isin"] in covered_sell_isins:
                continue

        txs.append({
            "action": "sell" if row["is_closed"] else "buy",
            "date": row["sell_date"] if row["is_closed"] else row["buy_date"],
            "isin": row["isin"],
            "nom": row["nom"],
            "tipus": row["tipus"],
            "custodian": row["custodian"],
            "units": row["units"],
            "nav": row["nav"],
            "valueEur": row["valueEur"],
        })
    return txs


def serialize_transactions(drafts: list[PMTransactionDraft]) -> list[PMTransactionRow]:
    records: list[PMTransactionRow] = []
    for tx in drafts:
        records.append({
            "action": tx["action"],
            "date": tx["date"].isoformat() if isinstance(tx["date"], date) else None,
            "isin": tx["isin"],
            "nom": tx["nom"],
            "tipus": tx["tipus"],
            "custodian": tx["custodian"],
            "units": tx["units"],
            "nav": round(tx["nav"], 4) if tx["nav"] is not None else None,
            "valueEur": round(tx["valueEur"], 0) if tx["valueEur"] is not None else None,
            "id": tx["id"],
        })
    return records


def main():
    wb = load_wb()
    print("Loading ETF price CSVs…")
    etf_prices = load_etf_prices()
    print(f"  {len(etf_prices)} ISINs with price history")
    bank_movements = load_bank_movements()
    print(f"Loaded {len(bank_movements)} resolved bank movements")

    print("Reading ETF tranches…")
    etf = etf_buys(wb, etf_prices)
    print(f"  {len(etf)} ETF buy tranches")

    print("Reading Master sheet positions…")
    master_rows = master_positions(wb)
    bank_exact, covered_buy_isins, covered_sell_isins = bank_txs(master_rows, bank_movements)
    master = fallback_master_txs(master_rows, covered_buy_isins, covered_sell_isins)
    buys  = [t for t in master if t["action"] == "buy"]
    sells = [t for t in master if t["action"] == "sell"]
    exact_buys = [t for t in bank_exact if t["action"] == "buy"]
    exact_sells = [t for t in bank_exact if t["action"] == "sell"]
    print(f"  {len(exact_buys)} exact bank buys / onboarding transfers")
    print(f"  {len(exact_sells)} exact bank sells")
    print(f"  {len(buys)} fallback managed-fund buys")
    print(f"  {len(sells)} fallback managed-fund sells (TANCATS)")

    draft_txs = etf + bank_exact + master
    # Sort by date (None dates last)
    draft_txs.sort(key=lambda t: t["date"] or date(2099, 1, 1))

    # Assign stable IDs
    for i, tx in enumerate(draft_txs):
        tx["id"] = f"tx-{i:04d}"
    all_txs = serialize_transactions(draft_txs)

    payload = json.dumps(all_txs, ensure_ascii=False, indent=None, separators=(",", ":"))
    out = (
        "// Auto-generated by scripts/transactions_export_js.py — do not edit manually\n"
        f"export const PM_TRANSACTIONS = {payload};\n"
    )
    OUT_JS.parent.mkdir(parents=True, exist_ok=True)
    OUT_JS.write_text(out, encoding="utf-8")
    print(f"\nTotal transactions: {len(all_txs)}")
    print(f"Output: {OUT_JS}")

    # Preview first/last 3
    print("\nSample (first 3):")
    for tx in all_txs[:3]:
        print(f"  {tx['date']} {tx['action']:4} {tx['isin']} {tx['nom'][:40]}")
    print("Sample (last 3):")
    for tx in all_txs[-3:]:
        print(f"  {tx['date']} {tx['action']:4} {tx['isin']} {tx['nom'][:40]}")


if __name__ == "__main__":
    main()
